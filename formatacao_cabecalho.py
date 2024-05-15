import pandas as pd

from datetime import time, datetime

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, NamedStyle, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def alunosLista():
    tabela = pd.read_excel("./bases/Lista geral.xlsx") # Base de dados geral.
    tabela.loc[:, 'ASSINATURA'] = '' # Criando campo de assinatura.
    dia_escolhido = "SEG" #Escolhendo dia para a lista.
    #Escolhendo o turno que deve ser impresso
    turno = 0  #matutino = 1 vespertino = 0

    dia = tabela.loc[tabela['DIA']==dia_escolhido] # Buscando com base no dia.

    #Filtrando por turno:

    if (turno):
        dados_lista = dia[dia['HORÁRIO ÍNICIO'].apply(lambda x: time(6, 0) <= x <= time(11, 59))]
    else:
        dados_lista = dia[dia['HORÁRIO ÍNICIO'].apply(lambda x: time(12, 0) <= x <= time(22, 59))]

    
    dados_lista = dados_lista.drop('DIA', axis=1) #Excluindo a coluna dias que vem junto com a base.
    #dados_lista.to_excel("./bases/ListaAssinatura.xlsx", index=False, sheet_name='Assinar')
    
    return dados_lista #Retornando a lista

def mesclarPlanilhas():
    #Formatando cabeçalho:
    pasta_cabecaclho = openpyxl.Workbook()
    cabecalho = pasta_cabecaclho.active
    # Inserir valores nas células específicas
    cabecalho['A1'] = "LISTA DE CHAMADA - INTERATIVO VESPERTINO"
    cabecalho['A2'] = "DIA"
    cabecalho['B2'] = "SEGUNDA"
    cabecalho['E2'] = "DATA AULA:  ____/_____/________"
    cabecalho['D3'] = "ORIENTADOR:"

    cabecalho['A4'] = "HORA INICIO"
    cabecalho['B4'] = "HORA FIM"
    cabecalho['C4'] = "CURSO"
    cabecalho['D4'] = "NOME"
    cabecalho['E4'] = "ASSINATURA:"

    original = openpyxl.Workbook()

    planilha_assinatura = alunosLista()
    planilha_cabecalho = cabecalho

    mescla = original.active
    # Copia os dados da planilha de cabeçalho para a planilha mescla
    for row_index, row in enumerate(planilha_cabecalho.iter_rows(), start=1):
        for cell_index, cell in enumerate(row, start=1):
            mescla.cell(row=row_index, column=cell_index).value = cell.value

    # Obtém o número total de linhas na planilha mescla
    num_linhas_mescla = mescla.max_row

    # Copia os dados da planilha de assinatura para a planilha mescla, começando da linha seguinte à última linha preenchida na mescla
    for row_index, (index, row)in enumerate(planilha_assinatura.iterrows(), start=num_linhas_mescla+1):
        for cell_index, cell in enumerate(row, start=1):
            mescla.cell(row=row_index, column=cell_index).value = cell
    
    #Meclando titulo
    mescla.merge_cells('A1:E1')

    #Definindo largura e altura das colunas e da linha do tituo:
    mescla.column_dimensions['A'].width = 11
    mescla.column_dimensions['B'].width = 11
    mescla.column_dimensions['C'].width = 13
    mescla.column_dimensions['D'].width = 41
    mescla.column_dimensions['E'].width = 51
    mescla.row_dimensions[4].height = 31

    # Criando borda simples

    borda = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    # Alinhamentos centralizar

    alinhamento = Alignment(horizontal='center', vertical='center', wrap_text=True)

    #Fonte em negrito
    fonte_negrito = Font(bold=True)
    format_titulo_cols = Font(size=12, bold=True)

    for row in range(4, mescla.max_row + 1):
            
        for col in range(1, 6):
            mescla.cell(row=row, column=col).border = borda
            if (row == 4):
                mescla.cell(row=row, column=col).alignment = alinhamento
                mescla.cell(row=row, column=col).font = format_titulo_cols
            if (col == 1 or col == 2) and row > 4:
                valor_celula = mescla.cell(row=row, column=col).value
                if valor_celula:
                    mescla.cell(row=row, column=col).alignment = alinhamento

    #Editando cabeçalho
    mescla.cell(row=1, column=1).alignment = alinhamento
    mescla.cell(row=1, column=1).font = fonte_negrito
    mescla.cell(row=2, column=1).font = fonte_negrito
    mescla.cell(row=2, column=2).font = fonte_negrito
    mescla.cell(row=2, column=5).font = fonte_negrito
    mescla.cell(row=3, column=4).font = fonte_negrito
    mescla.cell(row=1, column=1).font = Font(size=18)

    # Salva as alterações no arquivo
    original.save('./bases/ListaAssinatura.xlsx')