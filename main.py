import pandas as pd
from datetime import time
import openpyxl

from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

from openpyxl.styles import Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border

# Base de dados geral.

tabela = pd.read_excel("./bases/Lista geral.xlsx")
tabela.loc[:, 'ASSINATURA'] = ''

# Dados para buscar:

# Buscando por dia.

dia_escolhido = "SEG"

#matutino = 1 vespertino = 0

turno = 0

# Buscando com base nas informações solicitadas.

#Buscando por dia
dia = tabela.loc[tabela['DIA']==dia_escolhido]

#Buscando por turno:

if (turno):
    dados_lista = dia[dia['HORÁRIO ÍNICIO'].apply(lambda x: time(6, 0) <= x <= time(11, 59))]
else:
    dados_lista = dia[dia['HORÁRIO ÍNICIO'].apply(lambda x: time(12, 0) <= x <= time(22, 59))]

#Excluindo a coluna dias

dados_lista = dados_lista.drop('DIA', axis=1)

#Criando a lista

dados_lista.to_excel("./bases/ListaAssinatura.xlsx", index=False, sheet_name='Assinar')

#Formatando cabeçalho:

workbook = openpyxl.Workbook()

cabecalho = workbook.active

# Inserir valores nas células específicas
cabecalho['A1'] = "LISTA DE CHAMADA - INTERATIVO VESPERTINO"
cabecalho['A2'] = "DIA"
cabecalho['B2'] = "SEGUNDA"
cabecalho['E2'] = "DATA AULA: ____/_____/________"
cabecalho['D3'] = "ORIENTADOR:"

workbook.save('./bases/cabecalho.xlsx')

original = openpyxl.Workbook()
mescla = original.active

lista_assinatura = openpyxl.load_workbook('./bases/ListaAssinatura.xlsx')
cabecalho = openpyxl.load_workbook('./bases/cabecalho.xlsx')

planilha_assinatura = lista_assinatura['Assinar']
planilha_cabecalho = cabecalho['Sheet']

# Copia os dados da planilha de cabeçalho para a planilha mescla
for row_index, row in enumerate(planilha_cabecalho.iter_rows(), start=1):
    for cell_index, cell in enumerate(row, start=1):
        mescla.cell(row=row_index, column=cell_index).value = cell.value

# Obtém o número total de linhas na planilha mescla
num_linhas_mescla = mescla.max_row

# Copia os dados da planilha de assinatura para a planilha mescla, começando da linha seguinte à última linha preenchida na mescla
for row_index, row in enumerate(planilha_assinatura.iter_rows(), start=num_linhas_mescla+1):
    for cell_index, cell in enumerate(row, start=1):
        mescla.cell(row=row_index, column=cell_index).value = cell.value

#Formatando planilha:

#Meclando titulo
mescla.merge_cells('A1:E1')

#Definindo largura e altura das colunas e da linha do tituo:
mescla.column_dimensions['A'].width = 11
mescla.column_dimensions['B'].width = 11
mescla.column_dimensions['C'].width = 13
mescla.column_dimensions['D'].width = 41
mescla.column_dimensions['E'].width = 51
mescla.row_dimensions[4].height = 31

# Criando borda simples

borda = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))

# Alinhamentos centralizar

alinhamento = Alignment(horizontal='center', vertical='center', wrap_text=True)

#Fonte em negrito
fonte_negrito = Font(bold=True)
format_titulo_cols = Font(size=12, bold=True)

for row in range(4, mescla.max_row + 1):
        
    for col in range(1, 6):
        mescla.cell(row=row, column=col).border = borda
        if (row == 4):
            mescla.cell(row=row, column=col).alignment = alinhamento
            mescla.cell(row=row, column=col).font = format_titulo_cols
        if (col == 1 or col == 2) and row > 4:
            valor_celula = mescla.cell(row=row, column=col).value
            if valor_celula:
                partes = valor_celula.split(':')[:2]
                valor_formatado = ':'.join(partes)    # Juntar as partes de volta usando ':'
                mescla.cell(row=row, column=col).value = valor_formatado #Aplicando na célula
                mescla.cell(row=row, column=col).alignment = alinhamento

#Editando cabeçalho
mescla.cell(row=1, column=1).alignment = alinhamento
mescla.cell(row=1, column=1).font = fonte_negrito
mescla.cell(row=2, column=1).font = fonte_negrito
mescla.cell(row=2, column=2).font = fonte_negrito
mescla.cell(row=2, column=5).font = fonte_negrito
mescla.cell(row=3, column=4).font = fonte_negrito
mescla.cell(row=1, column=1).font = Font(size=18)


# Salva as alterações no arquivo
original.save('./bases/ListaAssinatura.xlsx')